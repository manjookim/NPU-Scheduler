#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
mz3 무설정(기본값) 실험 결과 CSV 2종(Set A: fps=0 / Set B: fps=30)을
기존 실험 xlsx와 같은 3시트 구조로 합친다.
  시트1 컬럼설명 / 시트2 전체(42행) / 시트3 케이스별_3회평균(14행)

사용: python build_xlsx_mz3.py
"""
import csv
import os
import statistics as st

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
CSV_DIR = os.path.join(HERE, "csv")
SET_A = os.path.join(CSV_DIR, "results_mz3_default.csv")        # fps=0
SET_B = os.path.join(CSV_DIR, "results_mz3_default_fps30.csv")  # fps=30
OUT = os.path.join(CSV_DIR, "results_mz3_default_combined.xlsx")

COND_ORDER = ["ssd", "deeplab", "mnv2", "ssd_deeplab", "ssd_mnv2",
              "deeplab_mnv2", "ssd_deeplab_mnv2"]

COL_DESC = [
    ("실험 조건",
     "Hailo Model Zoo 3모델(ssd_mobilenet_v1 / deeplab_v3_mobilenet_v2_wo_dilation / mobilenet_v2_1.0)을 "
     "Hailo-8L(npu-rpi1)에서 동시 스케줄링. 스케줄러 파라미터(priority/threshold/timeout/batch_size/power_mode)를 "
     "일절 설정하지 않은 HailoRT 기본값 조건. 단독 3 + 쌍 3 + 트리플 1 = 7조건 x 3회 반복 = 21회를 "
     "입력속도 2종(fps=0 무제한 / fps=30 제한)으로 각각 수행 = 총 42회."),
    ("input_fps",
     "writer 스레드가 프레임을 밀어넣는 속도(FPS). 0 = 무제한(큐가 포화될 때까지 최대속도). "
     "30 = 모델당 초당 30장으로 제한 — 큐 대기 성분을 줄여 순수 스케줄링 지연을 보기 위한 조건."),
    ("tag", "워크로드 조합 라벨. ssd / deeplab / mnv2 / ssd_deeplab / ssd_mnv2 / deeplab_mnv2 / ssd_deeplab_mnv2"),
    ("run_id", "같은 조건에서 몇 번째 반복인지 (1~3)"),
    ("frames", "모델당 처리한 프레임 수 (sampled_val2017 673장 전량)"),
    ("use_ssd / use_deeplab / use_mnv2", "이 실행에 해당 모델이 활성화됐는지 (1=사용, 0=미사용)"),
    ("ssd_latency_ms / deeplab_latency_ms / mnv2_latency_ms",
     "모델별 평균 latency(ms). enqueue(write 직전)~dequeue(모든 출력 read 완료) 구간. "
     "기존 hailo_cpp_test/model_runner.hpp 와 동일한 계측 지점 — 이전 실험과 비교 가능. "
     "비활성 모델은 NaN. [주의] input_fps=0에서는 큐 대기시간이 포함된 값임."),
    ("ssd_fps / deeplab_fps / mnv2_fps", "모델별 실측 처리량 = 처리 프레임 수 / 해당 모델 총 소요시간"),
    ("ssd_preprocess_ms / deeplab_preprocess_ms / mnv2_preprocess_ms",
     "장당 평균 전처리 시간(ms) = imread + resize(모델 입력 크기) + BGR2RGB. "
     "[주의] 이 3모델은 YOLO 계열이 아니라 letterbox가 아닌 단순 resize를 쓴다(Model Zoo 전처리 기준)."),
    ("ssd_total_time_s / deeplab_total_time_s / mnv2_total_time_s", "해당 모델이 전체 프레임을 처리하는 데 걸린 시간(초)"),
    ("cpu_percent", "실행 구간 평균 CPU 사용률(%) — /proc/stat 기반"),
    ("mem_percent", "실행 구간 평균 메모리 사용률(%) — /proc/meminfo 기반"),
    ("voluntary_ctx_switches / nonvoluntary_ctx_switches",
     "활성 모델들의 writer/reader 스레드 컨텍스트 스위치 합계 (/proc/thread-self/status)"),
    ("wall_time_s", "이 실행 전체의 wall-clock 시간(초)"),
    ("n_runs (평균 시트만)", "이 조건에서 평균낸 반복 횟수 (전부 3)"),
]


def load(path, fps_label):
    if not os.path.exists(path):
        print(f"  [건너뜀] 파일 없음: {path}")
        return []
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    print(f"  [로드] {os.path.basename(path)} — {len(rows)}행 (input_fps={fps_label})")
    return rows


def as_num(v):
    if v is None:
        return "NaN"
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return "NaN"
    try:
        return float(s)
    except ValueError:
        return s


def style_header(ws, ncols):
    fill = PatternFill("solid", fgColor="DDEBF7")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def main():
    print("=== mz3 결과 xlsx 생성 ===")
    rows_a = load(SET_A, 0)
    rows_b = load(SET_B, 30)
    rows = rows_a + rows_b
    if not rows:
        raise SystemExit("CSV가 하나도 없습니다.")

    headers = list(rows[0].keys())
    wb = openpyxl.Workbook()

    # ── 시트1: 컬럼설명 ──
    ws = wb.active
    ws.title = "컬럼설명"
    ws.append(["컬럼명", "설명"])
    for name, desc in COL_DESC:
        ws.append([name, desc])
    style_header(ws, 2)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 118
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # ── 시트2: 전체 ──
    ws2 = wb.create_sheet(f"전체({len(rows)}행)")
    ws2.append(headers)
    for r in rows:
        ws2.append([as_num(r[h]) for h in headers])
    style_header(ws2, len(headers))
    for i, h in enumerate(headers, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = max(11, min(len(h) + 3, 26))

    # ── 시트3: 조건별 3회 평균 ──
    numeric = [h for h in headers if h not in ("tag", "run_id")]
    groups = {}
    for r in rows:
        groups.setdefault((r["tag"], r["input_fps"]), []).append(r)

    ws3 = wb.create_sheet(f"케이스별_3회평균({len(groups)}행)")
    ws3.append(["n_runs", "tag"] + numeric)
    for fps in ("0", "30"):
        for tag in COND_ORDER:
            g = groups.get((tag, fps))
            if not g:
                continue
            out = [len(g), tag]
            for h in numeric:
                vals = []
                for r in g:
                    v = as_num(r[h])
                    if isinstance(v, float):
                        vals.append(v)
                out.append(round(st.mean(vals), 4) if vals else "NaN")
            ws3.append(out)
    style_header(ws3, len(numeric) + 2)
    ws3.column_dimensions["B"].width = 20
    for i in range(3, len(numeric) + 3):
        ws3.column_dimensions[get_column_letter(i)].width = 14

    wb.save(OUT)
    print(f"  -> 저장: {OUT}")
    print(f"     전체 {len(rows)}행 / 평균 {len(groups)}조건")


if __name__ == "__main__":
    main()
