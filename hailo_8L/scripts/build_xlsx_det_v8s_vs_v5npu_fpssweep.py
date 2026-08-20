#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_xlsx_det_v8s_vs_v5npu_fpssweep.py — (Hailo-8L, rpi1) Det 단일모델
v8s_h8l(CPU 후처리) vs v5-nms_core(NPU 후처리) x INPUT_FPS{0,30} 실험 결과를
컬럼설명/전체/조건별_3회평균 3개 시트짜리 xlsx로 정리한다.

build_xlsx_det_v8s_vs_v5npu.py(FPS=60 고정판)의 후속 — input_fps 컬럼과 HRTT 유래
지표(npu_percent/switches_per_s/idle_time_pct/avg_fps_hrtt/avg_latency_hrtt/
max_latency_hrtt/activation_hrtt) 설명이 추가됐다.

사용법: python3 build_xlsx_det_v8s_vs_v5npu_fpssweep.py <raw.csv> <avg.csv> [출력.xlsx]
"""
import sys, csv
from openpyxl import Workbook
from openpyxl.styles import Alignment

COLUMN_DESC = [
    ("실험 조건",
     "Hailo-8L(rpi1). Det(Detection) 단일모델만 대상. HEF 3종 비교: yolov8s_h8l.hef(engine=cpu, "
     "후처리를 host CPU에서 수행) vs yolov5xs_wo_spp_nms_core.hef(hailo8l 타겟, engine=nn_core/"
     "auto, NMS를 포함한 후처리를 NPU neural core에서 수행 — hailortcli parse-hef로 output "
     "vstream이 이미 'HAILO NMS BY CLASS' 포맷으로 나옴을 실측 확인함) vs yolov5xs_wo_spp.hef "
     "([2026-08-20 추가] v5-nms_core와 동일 아키텍처(v5xs)에 '_nms_core' 접미사만 없는 버전, "
     "engine=cpu, 호스트 CPU에서 NMS까지 수행 — v5 아키텍처를 NPU/CPU 후처리 양쪽으로 비교하기 "
     "위해 추가). batch=1/threshold=1/timeout=0ms/priority=0으로 고정, INPUT_FPS만 {0(제한없음/"
     "최대속도), 30}로 스윕. 모델 버전(v8s/v5)과 입력 크기(640/512)는 YOLOv8이 NPU 후처리를 "
     "지원하지 않아 불가피하게 다름(PROJECT_SUMMARY.md §6) — 그 외 조건은 모두 통일."),
    ("run_id", "전체 실행 순번(1~18) — FPS=0 구간(1~6: v5-NPU 3회 → v8s-CPU 3회) → FPS=30 구간"
               "(7~12: v5-NPU 3회 → v8s-CPU 3회) → [2026-08-20 추가] v5-CPU 3회(FPS=0, 13~15) → "
               "v5-CPU 3회(FPS=30, 16~18) 순서"),
    ("hef_name", "실행된 모델 라벨 — 'Detection-CPU-baseline'(yolov8s_h8l, CPU 후처리), "
                 "'YOLOv5-NPU-postprocess'(yolov5xs_wo_spp_nms_core, NPU 후처리), 또는 "
                 "'YOLOv5-CPU-baseline'(yolov5xs_wo_spp, CPU 후처리)"),
    ("img_size", "모델 입력 한 변 크기(정사각) — v8s_h8l=640, v5xs_nms_core/v5xs_wo_spp=512"),
    ("input_fps", "입력 프레임 페이싱 목표치. 0=제한 없음(프레임 간 sleep 없이 큐가 받는 대로 "
                  "최대 속도 투입), 30=33ms 간격으로 페이싱(카메라 30fps 실입력 모사). "
                  "0/0.0 이 아니라 정말로 '무제한'이라는 뜻이며, model_runner.hpp의 "
                  "if(INPUT_FPS>0){sleep} 로직으로 구현됨."),
    ("batch / threshold / timeout_ms / priority",
     "HailoRT 스케줄러 파라미터. 이 실험에서는 1/1/0/0으로 모든 조건 동일 고정(모델 1개 단독 "
     "실행이라 스케줄링 경합 자체가 없음)"),
    ("frame_count", "정상적으로 latency가 측정된 프레임 수(데이터셋 전체, 673장)"),
    ("avg_preprocess_ms", "장당 평균 전처리 시간(ms) — imread + letterbox + BGR2RGB"),
    ("avg_latency_ms", "장당 평균 추론 시간(ms) — 입력 write ~ 출력 read 왕복(enqueue~dequeue), "
                       "코드가 자체 chrono 타이머로 직접 측정"),
    ("avg_postprocess_ms",
     "[요청하신 '후처리 시간 별도 측정' 항목] 장당 평균 후처리 시간(ms) — decode_det()의 "
     "NMS-by-class 버퍼 파싱 비용. **주의(해석 시 반드시 참고)**: NPU 조건(YOLOv5-NPU-"
     "postprocess)은 NMS 연산 자체가 이미 칩에서 끝나고 나온 결과를 '파싱'만 하는 시간이고, "
     "CPU 조건(Detection-CPU-baseline, YOLOv5-CPU-baseline)은 host에서 NMS 연산까지 전부 수행한 "
     "시간이다 — 조건 간 이 컬럼 값을 곧이곧대로 비교하면 'NPU가 압도적으로 빠르다'는 착시가 생길 수 "
     "있음(실제로는 서로 다른 작업량을 재는 것). 후처리 부담이 어느 프로세서로 갔는지는 "
     "avg_latency_ms(추론 시간에 NMS가 얼마나 녹아들었는지) / cpu_percent(호스트 CPU 점유율) "
     "/ npu_percent를 같이 봐야 공정하게 비교 가능."),
    ("avg_total_time_ms", "장당 전체 시간(ms) = avg_preprocess_ms + avg_latency_ms + avg_postprocess_ms"),
    ("total_time_s", "이 조건에서 데이터셋 전체(673장)를 처리하는 데 걸린 총 wall-time(초)"),
    ("run_time_s", "프로그램이 이 실행에 소요한 전체 시간(초) — VDevice 생성 이후 ~ 추론 종료까지"),
    ("cpu_percent / mem_percent",
     "실행 중 host CPU/메모리 사용률(%) — NPU 조건이 CPU 조건보다 낮게 나오면 '후처리 부담이 "
     "실제로 host CPU에서 NPU로 옮겨갔다'는 방향성 증거로 참고 가능"),
    ("voluntary_ctx_switches / nonvoluntary_ctx_switches", "writer+reader 스레드의 컨텍스트 스위치 횟수 합"),
    ("npu_percent",
     "실행 스크립트가 hailo_utilization.py(별도 프로세스, /tmp/hmon_files 기반 모니터링)로 "
     "실시간 측정한 NPU 칩 사용률 평균(%) — chrono 타이머와 무관한 별도 측정 경로"),
    ("switches_per_s / idle_time_pct",
     "HRTT(HAILO_MONITOR=1 트레이스) 파싱 결과 — 초당 스케줄러 컨텍스트 스위치 횟수 / 전체 "
     "구간 중 아무 모델도 activate 안 된 idle 비율(%). 단일모델 실험이라 switch는 원래 거의 "
     "0에 가깝고, idle_time_pct도 낮게 나오는 게 정상(경쟁 모델이 없어 스케줄링 전환이 "
     "일어날 이유가 적음)."),
    ("avg_fps_hrtt / avg_latency_hrtt / max_latency_hrtt / activation_hrtt",
     "HRTT 트레이스에서 직접 계산한 값 — avg_fps_hrtt(D2H 프레임 dequeue 간격 기반 FPS), "
     "avg_latency_hrtt/max_latency_hrtt(H2D→D2H 페어링 기반 평균/최대 latency, ms), "
     "activation_hrtt(모델이 디바이스에 activate돼 있던 평균 시간, ms). avg_latency_ms(코드 "
     "자체 chrono 측정)와 avg_latency_hrtt(HRTT 기반 측정)는 측정 방식이 달라 수치가 다를 수 "
     "있음 — 서로 다른 관측점에서 잰 값이라는 점을 감안해서 비교할 것."),
    ("n_runs (평균 시트에만)", "그룹에 포함된 반복 횟수(=3)"),
]


def read_csv_rows(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    raw_csv, avg_csv = sys.argv[1], sys.argv[2]
    out_path = sys.argv[3] if len(sys.argv) > 3 else "results_det_v8s_vs_v5npu_fpssweep_combined.xlsx"

    raw_rows = read_csv_rows(raw_csv)
    avg_rows = read_csv_rows(avg_csv)

    wb = Workbook()
    ws_desc = wb.active
    ws_desc.title = "컬럼설명"
    ws_desc.append(["컬럼명", "설명"])
    for name, desc in COLUMN_DESC:
        ws_desc.append([name, desc])
    ws_desc.column_dimensions["A"].width = 32
    ws_desc.column_dimensions["B"].width = 110
    for row in ws_desc.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    ws_all = wb.create_sheet(f"전체({len(raw_rows) - 1}행)")
    for row in raw_rows:
        ws_all.append(row)

    ws_avg = wb.create_sheet(f"조건별_3회평균({len(avg_rows) - 1}행)")
    for row in avg_rows:
        ws_avg.append(row)

    wb.save(out_path)
    print(f"완료: {out_path}")


if __name__ == "__main__":
    main()
