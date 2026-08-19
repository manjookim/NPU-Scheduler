#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_xlsx_det_v8s_vs_v5npu.py — (Hailo-8L, rpi1) Det 단일모델 v8s_h8l(CPU 후처리) vs
v5-nms_core(NPU 후처리) FPS=60 실험 결과를 컬럼설명/전체/조건별_3회평균 3개 시트짜리
xlsx로 정리한다. (results_singlemodel_fps5to25.xlsx와 동일한 시트 구성 관례)

PC(또는 openpyxl이 설치된 아무 환경)에서 실행하면 된다. RPi에서 만들 필요 없음 —
raw CSV와 avg CSV를 scp로 받아온 뒤 로컬에서 돌리는 용도.

사용법: python3 build_xlsx_det_v8s_vs_v5npu.py <raw.csv> <avg.csv> [출력.xlsx]
"""
import sys, csv
from openpyxl import Workbook

COLUMN_DESC = [
    ("실험 조건",
     "Hailo-8L(rpi1). Det(Detection) 단일모델만 대상. HEF 2종 비교: yolov8s_h8l.hef(engine=cpu, "
     "후처리를 host CPU에서 수행) vs yolov5xs_wo_spp_nms_core.hef(hailo8l 타겟, engine=nn_core/"
     "auto, 후처리 상당 부분을 NPU neural core에서 수행). batch=1/threshold=1/timeout=0ms/"
     "priority=0/INPUT_FPS=60으로 두 조건 동일 고정. 모델 버전(v8s/v5)과 입력 크기(640/512)는 "
     "YOLOv8이 NPU 후처리를 지원하지 않아 불가피하게 다름(PROJECT_SUMMARY.md §6) — 그 외 "
     "조건은 모두 통일."),
    ("run_id", "같은 조건 안에서 몇 번째 반복 실행인지"),
    ("hef_name", "실행된 모델 라벨 — 'Detection-CPU-baseline'(yolov8s_h8l, CPU 후처리) 또는 "
                 "'YOLOv5-NPU-postprocess'(yolov5xs_wo_spp_nms_core, NPU 후처리)"),
    ("img_size", "모델 입력 한 변 크기(정사각) — v8s_h8l=640, v5xs_nms_core=512"),
    ("batch / threshold / timeout_ms / priority",
     "HailoRT 스케줄러 파라미터. 이 실험에서는 1/1/0/0으로 두 조건 동일 고정(모델 1개 단독 "
     "실행이라 스케줄링 경합 자체가 없음)"),
    ("frame_count", "정상적으로 latency가 측정된 프레임 수(데이터셋 전체, 보통 약 600여장)"),
    ("avg_preprocess_ms", "장당 평균 전처리 시간(ms) — imread + letterbox + BGR2RGB"),
    ("avg_latency_ms", "장당 평균 추론 시간(ms) — 입력 write ~ 출력 read 왕복(enqueue~dequeue)"),
    ("avg_postprocess_ms",
     "장당 평균 후처리 시간(ms) — decode_det()의 NMS-by-class 버퍼 파싱 비용. [주의] NPU 조건은 "
     "NMS 연산 자체가 이미 칩에서 끝나고 나온 결과를 '파싱'만 하는 시간이고, CPU 조건은 host에서 "
     "NMS까지 수행한 시간이라 이 컬럼만으로 직접 비교하면 왜곡될 수 있음 — avg_latency_ms / "
     "avg_total_time_ms / cpu_percent를 같이 볼 것"),
    ("avg_total_time_ms", "장당 전체 시간(ms) = avg_preprocess_ms + avg_latency_ms + avg_postprocess_ms"),
    ("total_time_s", "이 조건에서 데이터셋 전체를 처리하는 데 걸린 총 wall-time(초)"),
    ("run_time_s", "프로그램이 이 실행에 소요한 전체 시간(초) — VDevice 생성 이후 ~ 추론 종료까지"),
    ("cpu_percent / mem_percent",
     "실행 중 host CPU/메모리 사용률(%) — NPU 조건이 CPU 조건보다 낮게 나오면 '후처리 부담이 "
     "실제로 host CPU에서 NPU로 옮겨갔다'는 방향성 증거로 참고 가능"),
    ("voluntary_ctx_switches / nonvoluntary_ctx_switches", "writer+reader 스레드의 컨텍스트 스위치 횟수 합"),
    ("n_runs (평균 시트에만)", "그룹에 포함된 반복 횟수(=3)"),
]


def read_csv_rows(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    raw_csv, avg_csv = sys.argv[1], sys.argv[2]
    out_path = sys.argv[3] if len(sys.argv) > 3 else "results_det_v8s_vs_v5npu_fps60_combined.xlsx"

    raw_rows = read_csv_rows(raw_csv)
    avg_rows = read_csv_rows(avg_csv)

    wb = Workbook()
    ws_desc = wb.active
    ws_desc.title = "컬럼설명"
    ws_desc.append(["컬럼명", "설명"])
    for name, desc in COLUMN_DESC:
        ws_desc.append([name, desc])
    ws_desc.column_dimensions["A"].width = 32
    ws_desc.column_dimensions["B"].width = 110

    ws_all = wb.create_sheet(f"전체({len(raw_rows) - 1}행)")
    for row in raw_rows:
        ws_all.append(row)

    ws_avg = wb.create_sheet(f"조건별_3회평균({len(avg_rows) - 1}행)")
    for row in avg_rows:
        ws_avg.append(row)

    wb.save(out_path)
    print(f"완료: {out_path}")


if __name__ == "__main__":
    main()
