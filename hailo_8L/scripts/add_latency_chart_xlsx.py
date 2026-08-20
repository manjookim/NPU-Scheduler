#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
add_latency_chart_xlsx.py
results_det_v8s_vs_v5npu_fpssweep_combined.xlsx에 "그래프" 시트를 추가하고, 엑셀
네이티브 막대그래프 2개를 삽입한다(조교님께 전달할 때 엑셀에서 바로 보이고, 값도
클릭하면 수정 가능한 진짜 차트 객체로 남기기 위함 — 이미지 붙여넣기 아님).

  차트 A: 조건별(hef_name x input_fps) 3회 평균 latency 비교
  차트 B: 조건별 개별 실행(run 1~3) latency 비교 — "각 케이스 테스트별" 요청 반영

[2026-08-20] 하드코딩된 값(MEAN_ROWS/RUN_ROWS) 대신 raw.csv/avg.csv에서 직접 읽어
동적으로 채우도록 변경 — v5-CPU(YOLOv5-CPU-baseline, wo_spp, engine=cpu) 조건이
새로 추가되면서 모델이 2종에서 3종으로 늘어나 하드코딩 유지가 비현실적이라 반영.

색상은 hailo_8L 실험 시리즈에서 쓰는 카테고리 팔레트와 동일하게 고정(모델 정체성은
색상, FPS는 명도로 구분):
  v5-NPU 후처리 = blue(#2a78d6/#86b6ef), v8s-CPU 후처리 = orange(#eb6834/#f3a679),
  v5-CPU 후처리 = green(#2ca02c/#8fd18f) [2026-08-20 추가]

사용법: python add_latency_chart_xlsx.py <xlsx_path> <raw_csv_path> <avg_csv_path>
       (같은 xlsx 파일에 "그래프" 시트를 추가/교체 후 저장)
"""
import sys, csv
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# hef_name -> (짧은 라벨, dark(FPS0) 색상, light(FPS30) 색상). 순서 = 차트에 표시되는 순서.
MODEL_INFO = [
    ("YOLOv5-NPU-postprocess", "v5-NPU",  "2A78D6", "86B6EF"),
    ("Detection-CPU-baseline", "v8s-CPU", "EB6834", "F3A679"),
    ("YOLOv5-CPU-baseline",    "v5-CPU",  "2CA02C", "8FD18F"),
]


def style_series(series, hex_color):
    series.graphicalProperties.solidFill = hex_color
    series.graphicalProperties.line.noFill = True


def read_csv_dicts(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    xlsx_path, raw_csv_path, avg_csv_path = sys.argv[1], sys.argv[2], sys.argv[3]

    avg_rows = read_csv_dicts(avg_csv_path)
    raw_rows = read_csv_dicts(raw_csv_path)

    models = [m for m in MODEL_INFO if any(r["hef_name"] == m[0] for r in avg_rows)]
    fps_list = sorted({r["input_fps"] for r in avg_rows}, key=lambda x: float(x))

    # (hef_name, input_fps) -> avg_latency_ms(평균)
    avg_lookup = {(r["hef_name"], r["input_fps"]): float(r["avg_latency_ms"]) for r in avg_rows}
    # (hef_name, input_fps) -> [avg_latency_ms per run, run_id 순]
    raw_lookup = {}
    for r in raw_rows:
        key = (r["hef_name"], r["input_fps"])
        raw_lookup.setdefault(key, []).append(float(r["avg_latency_ms"]))

    wb = load_workbook(xlsx_path)
    if "그래프" in wb.sheetnames:
        del wb["그래프"]
    ws = wb.create_sheet("그래프")

    # ---------- 표 A: 조건별 평균 ----------
    ws["A1"] = "조건별 3회 평균 avg_latency_ms (ms)"
    ws["A2"] = ""
    for j, (hef_name, label, dark, light) in enumerate(models):
        ws.cell(row=2, column=2 + j, value=f"{label} ({hef_name})")
    for i, fps in enumerate(fps_list, start=3):
        ws.cell(row=i, column=1, value=f"FPS {fps}")
        for j, (hef_name, label, dark, light) in enumerate(models):
            ws.cell(row=i, column=2 + j, value=avg_lookup.get((hef_name, fps)))

    n_models = len(models)
    chart_a = BarChart()
    chart_a.type = "col"
    chart_a.grouping = "clustered"
    chart_a.title = "조건별 평균 Latency (ms)"
    chart_a.y_axis.title = "avg_latency_ms"
    chart_a.x_axis.title = "INPUT_FPS"
    chart_a.style = 2
    chart_a.height = 9
    chart_a.width = 16

    data_a = Reference(ws, min_col=2, max_col=1 + n_models, min_row=2, max_row=2 + len(fps_list))
    cats_a = Reference(ws, min_col=1, max_col=1, min_row=3, max_row=2 + len(fps_list))
    chart_a.add_data(data_a, titles_from_data=True)
    chart_a.set_categories(cats_a)
    for j, (hef_name, label, dark, light) in enumerate(models):
        style_series(chart_a.series[j], dark)
    chart_a.dataLabels = DataLabelList()
    chart_a.dataLabels.showVal = True
    ws.add_chart(chart_a, "A8")

    # ---------- 표 B: 조건별 개별 실행(run 1~3) ----------
    base_row = 8 + 18  # 차트 A 아래로 표 B 배치
    ws.cell(row=base_row - 1, column=1, value="조건별 개별 실행(run) avg_latency_ms (ms)")
    hdr_row = base_row
    ws.cell(row=hdr_row, column=1, value="")

    # 컬럼 순서: 모델별로 FPS0, FPS30 ... (기존 방식과 동일하게 모델이 바깥, FPS가 안쪽)
    # fps_list는 오름차순 정렬돼 있으므로 순서대로 dark(가장 낮은 FPS)->light로 명도를 밝혀간다.
    col_defs = []  # [(hef_name, fps, label, color)]
    for hef_name, label, dark, light in models:
        shades = [dark, light] if len(fps_list) == 2 else [dark] * len(fps_list)
        for fps, color in zip(fps_list, shades):
            col_defs.append((hef_name, fps, f"{label} · FPS{fps}", color))

    n_runs = max((len(v) for v in raw_lookup.values()), default=0)
    for j, (hef_name, fps, label, color) in enumerate(col_defs):
        ws.cell(row=hdr_row, column=2 + j, value=label)
    for i in range(n_runs):
        ws.cell(row=hdr_row + 1 + i, column=1, value=f"Run {i + 1}")
        for j, (hef_name, fps, label, color) in enumerate(col_defs):
            vals = raw_lookup.get((hef_name, fps), [])
            ws.cell(row=hdr_row + 1 + i, column=2 + j, value=vals[i] if i < len(vals) else None)

    chart_b = BarChart()
    chart_b.type = "col"
    chart_b.grouping = "clustered"
    chart_b.title = "케이스(반복 실행)별 Latency 비교 (ms)"
    chart_b.y_axis.title = "avg_latency_ms"
    chart_b.x_axis.title = "반복 실행"
    chart_b.style = 2
    chart_b.height = 9
    chart_b.width = 16

    last_row = hdr_row + n_runs
    data_b = Reference(ws, min_col=2, max_col=1 + len(col_defs), min_row=hdr_row, max_row=last_row)
    cats_b = Reference(ws, min_col=1, max_col=1, min_row=hdr_row + 1, max_row=last_row)
    chart_b.add_data(data_b, titles_from_data=True)
    chart_b.set_categories(cats_b)
    for j, (hef_name, fps, label, color) in enumerate(col_defs):
        style_series(chart_b.series[j], color)
    chart_b.dataLabels = DataLabelList()
    chart_b.dataLabels.showVal = True
    ws.add_chart(chart_b, f"A{last_row + 3}")

    col_widths = ["A"] + [chr(ord("B") + i) for i in range(max(n_models, len(col_defs)))]
    for col in col_widths:
        ws.column_dimensions[col].width = 22 if col != "A" else 20

    wb.save(xlsx_path)
    print(f"완료: '그래프' 시트(+차트 2개, 모델 {n_models}종) 추가 -> {xlsx_path}")


if __name__ == "__main__":
    main()
