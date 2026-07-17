#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""batch x priority 실험 종합 xlsx: 개요 + 3회평균 요약(전체/케이스별) + 분석 + 그래프 + 유사사례 + 원본"""
import os, glob, re, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_DIR = os.path.join(BASE, "results", "batch_priority")
AN = os.path.join(CSV_DIR, "analysis")
OUT = os.path.join(CSV_DIR, "results_batch_priority_SUMMARY.xlsx")

HDR = PatternFill("solid", fgColor="305496"); HDRF = Font(color="FFFFFF", bold=True)
TITLE = Font(bold=True, size=14); SUB = Font(bold=True, size=11, color="305496")

wb = Workbook(); wb.remove(wb.active)

def write_table(ws, rows, start=1, header=True):
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            c = ws.cell(row=start+i, column=1+j, value=v)
            if header and i == 0:
                c.fill = HDR; c.font = HDRF; c.alignment = Alignment(horizontal="center")
    # 열너비
    if rows:
        for j in range(len(rows[0])):
            w = max((len(str(r[j])) for r in rows if j < len(r)), default=10)
            ws.column_dimensions[get_column_letter(1+j)].width = min(max(w+2, 8), 22)

def load_csv(path):
    with open(path, encoding="utf-8") as f:
        return [row for row in csv.reader(f)]

# ---------- 00 개요 ----------
ws = wb.create_sheet("00_개요")
ws["A1"] = "batch × priority 실험 결과 요약 및 분석"; ws["A1"].font = TITLE
overview = [
 "",
 "[실험 개요]",
 "· Raspberry Pi 5 + Hailo-8L NPU, HailoRT 스케줄러(Round-Robin)",
 "· Detection / Segmentation / Pose 3개 모델을 NPU 공유하며 동시 추론",
 "· 변수: batch_size {1,10,50,63}(모델 통일), priority {0,15,31}(활성 모델별 교차)",
 "· 고정: threshold=1(기본값), timeout=0, 입력 sampled_val2017 673장, INPUT_FPS=0(최대 처리량)",
 "· 사용모델수(1/2/3) × batch(4) = 252 조건 × 3회 = 756회",
 "",
 "[시트 구성]",
 "· 01_요약_전체평균 : 252 조건 3회 평균 (한눈에)",
 "· avg_* (12개)     : 케이스별(사용모델수×batch) 3회 평균 시트",
 "· 02_batch효과 / 03_priority효과 / 04_상관관계 : 분석 표",
 "· 05_그래프        : batch 효과 · priority 효과 · 히트맵",
 "· 06_유사사례      : 관련 연구와 비교 분석",
 "· raw_* (12개)     : 원본 3회 데이터 (HRTT 컬럼 포함)",
 "",
 "[핵심 결론]",
 "· priority 효과: 우선순위 advantage(자기-최대타) 클수록 total_time 급감 (상관 ≈ -0.78, batch 무관)",
 "· batch 효과: 클수록 per-frame latency 급증(상관 ≈ +0.89), 그러나 batch=1은 처리량 낮아 total_time 최악",
 "            → batch≥10에서 total_time·fps 최적(포화), 이후는 latency만 악화 = 지연-처리량 트레이드오프",
 "· NPU 점유율은 전 구간 95~99%(포화). 우선순위 격차 크면 낮은 모델 starvation(총 처리시간 급증)",
]
for i, line in enumerate(overview, start=2):
    ws.cell(row=i, column=1, value=line)
    if line.startswith("["): ws.cell(row=i, column=1).font = SUB
ws.column_dimensions["A"].width = 95

# ---------- 01 요약 전체평균 ----------
ws = wb.create_sheet("01_요약_전체평균")
write_table(ws, load_csv(os.path.join(AN, "summary_all_avg.csv")))
ws.freeze_panes = "A2"

# ---------- avg_* 케이스별 ----------
for n in (1,2,3):
    for b in (1,10,50,63):
        p = os.path.join(AN, f"summary_{n}model_b{b}.csv")
        if not os.path.exists(p): continue
        ws = wb.create_sheet(f"avg_{n}m_b{b}")
        write_table(ws, load_csv(p)); ws.freeze_panes = "A2"

# ---------- 02~04 분석 ----------
for name, fn in [("02_batch효과","batch_effect.csv"),
                 ("03_priority효과","priority_effect.csv"),
                 ("04_상관관계","correlations.csv")]:
    ws = wb.create_sheet(name)
    write_table(ws, load_csv(os.path.join(AN, fn))); ws.freeze_panes = "A2"

# ---------- 05 그래프 ----------
ws = wb.create_sheet("05_그래프")
ws["A1"] = "분석 그래프"; ws["A1"].font = TITLE
imgs = [("A3","fig_batch_effect.png","batch 효과 (동일 우선순위)"),
        ("A40","fig_priority_effect.png","priority 효과 (advantage vs total_time)"),
        ("A70","fig_heatmap_det_b10.png","Det total_time 히트맵 (3모델 batch10)")]
r = 3
for _, fn, cap in imgs:
    path = os.path.join(AN, fn)
    if os.path.exists(path):
        ws.cell(row=r, column=1, value=cap).font = SUB
        im = XLImage(path); im.anchor = f"A{r+1}"
        ws.add_image(im); r += 34

# ---------- 06 유사사례 ----------
ws = wb.create_sheet("06_유사사례")
ws["A1"] = "유사 실험/연구 사례 비교"; ws["A1"].font = TITLE
lit = [
 ["연구/시스템","중점(요약한 것)","우리 실험과의 연결"],
 ["EdgeServing (Multi-DNN Serving)","런타임에 모델·exit·batch를 함께 선택해 시스템 전체 SLO 영향 최소화; 시간분할 GPU 공유","우리도 batch를 스케줄 knob로 봄 — batch가 latency·throughput을 동시에 좌우함을 정량 확인"],
 ["Miriam (Real-time Multi-DNN, Edge GPU)","co-running DNN의 latency-throughput 충돌을 elastic kernel로 완화; priority/criticality 반영","우리 결과의 지연-처리량 트레이드오프와 동일 문제의식; 우선순위로 완화되는 것을 관측"],
 ["Fluid Batching (Edge NPU)","batching(고활용)과 single-sample(저지연)의 장점을 preemptive 스케줄로 결합","우리: batch=1 저지연·저처리량 vs 큰 batch 고지연·고처리량 — 같은 트레이드오프의 양 끝"],
 ["ElasticRoom (Multi-tenant, strong priority)","강한 우선순위 스케줄로 실시간 요청 저지연+고활용 동시 달성 (Goodput 14~49%↑)","우리: priority advantage가 클수록 total_time 급감 — 우선순위의 실효성 정량화"],
 ["Fair/Firm RT Scheduling (RL, multi-accel.)","deadline hit/miss + 공정성 보상; starvation 방지","우리: 우선순위 격차 크면 낮은 모델 starvation(total_time 급증) — 공정성 부재 시 위험을 데이터로 확인"],
 ["SLO-aware priority + debt (multi-tenant)","지연 허용 워크로드를 throttle, debt로 starvation 방지","우리 threshold=0 hang이 곧 starvation 사례 — 실무적 anti-starvation 장치의 필요성 뒷받침"],
 ["Heterogeneous co-execution (mobile NPU/DSP)","동시 모델 수 증가 시 프로세서별 비균일 성능 저하(NPU +27%, DSP 붕괴)","우리: 1→2→3 모델로 갈수록 latency·total_time 비선형 증가 — 동시성 비용 확인"],
]
write_table(ws, lit)
ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 52; ws.column_dimensions["C"].width = 60
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment = Alignment(wrap_text=True, vertical="top")
note = ["","[요약: 관련 연구가 공통으로 중점 둔 것]",
 "1) 지연(특히 tail) vs 처리량(goodput) 트레이드오프",
 "2) 우선순위/시간분할로 실시간성 확보 + starvation·공정성 관리",
 "3) batch를 정적 상수가 아니라 스케줄 변수로 취급",
 "4) NPU/가속기 공유 시 동시성에 따른 비균일 성능 저하",
 "→ 우리 실험은 (1)(2)(3)을 batch·priority 2변수로 실측·정량화했고, threshold=0 starvation hang은 (2)의 필요성을 실증."]
r = ws.max_row + 2
for line in note:
    ws.cell(row=r, column=1, value=line)
    if line.startswith("["): ws.cell(row=r, column=1).font = SUB
    r += 1

# ---------- raw_* 원본 ----------
for n in (1,2,3):
    for b in (1,10,50,63):
        p = os.path.join(CSV_DIR, f"results_{n}model_b{b}.csv")
        if not os.path.exists(p): continue
        ws = wb.create_sheet(f"raw_{n}m_b{b}")
        write_table(ws, load_csv(p)); ws.freeze_panes = "A2"

wb.save(OUT)
print("저장:", OUT)
print("시트 수:", len(wb.sheetnames))
print("시트:", wb.sheetnames)
